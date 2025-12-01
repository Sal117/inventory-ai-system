# src/report_exporter.py
import pandas as pd
import io
import base64
import tempfile
import os
from datetime import datetime
from fpdf import FPDF
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
import numpy as np


from rapidfuzz import process, fuzz

from src.agent_tools import AgentTools
from src.ml.ml_predictor import MLPredictor
import streamlit as st

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# ------------------------------------------------------------
#  Section 1 req def etc...
# ------------------------------------------------------------

# ------------------------------------------------------------
#  GLOBAL STYLE PALETTE (Premium + Consistent)
# ------------------------------------------------------------
COLORS = {
    "primary": (40, 90, 190),        # Deep royal blue
    "accent": (80, 140, 255),        # Light blue
    "danger": (210, 50, 50),         # Red for warnings
    "success": (40, 170, 80),        # Green
    "text_dark": (30, 30, 30),
    "text_muted": (120, 120, 120),
    "border_light": (205, 205, 205),
    "border_subtle": (230, 230, 230),
    "table_alt": (245, 245, 245)
}


# ------------------------------------------------------------
#  HELPER — Convert Matplotlib Figure → Temp PNG Path (SAFE)
# ------------------------------------------------------------
def fig_to_png(fig):
    """
    Convert Matplotlib figure to a temporary PNG file path.
    Fully safe: never breaks the PDF even if figure is invalid.
    """
    try:
        tmp_path = os.path.join(
            tempfile.gettempdir(),
            f"chart_{datetime.now().timestamp()}.png"
        )
        fig.savefig(tmp_path, bbox_inches="tight", dpi=180)
        plt.close(fig)
        return tmp_path

    except Exception as e:
        print("❌ fig_to_png failed:", e)
        return None  # PDF section above can skip images safely

# ------------------------------------------------------------
#  PDF BASE CLASS WITH CLEAN HEADER/FOOTER + MARGINS
# ------------------------------------------------------------
class StyledPDF(FPDF):
    def __init__(self):
        super().__init__()
        self.set_auto_page_break(auto=True, margin=15)
        self.set_left_margin(12)
        self.set_right_margin(12)

    def header(self):
        try:
            self.set_font("DejaVu", "", 11)
            self.set_text_color(*COLORS["text_muted"])
            self.cell(
                0, 8,
                datetime.now().strftime("Generated on %Y-%m-%d %H:%M"),
                ln=True,
                align="R"
            )

            # subtle line under header
            self.set_draw_color(*COLORS["border_subtle"])
            self.set_line_width(0.2)
            self.line(10, self.get_y(), 200, self.get_y())
            self.ln(3)

        except:
            pass  # header should NEVER break PDF

    def footer(self):
        try:
            self.set_y(-15)
            self.set_font("DejaVu", "", 10)
            self.set_text_color(*COLORS["text_muted"])

            self.cell(0, 10, f"Page {self.page_no()}", align="C")
        except:
            pass

# ------------------------------------------------------------
#  MAIN EXPORTER CLASS
# ------------------------------------------------------------
class ReportExporter:

    def _new_pdf(self):
        """Creates a new styled PDF with fonts preloaded and safe defaults."""
        pdf = StyledPDF()

        font_dir = os.path.join(os.path.dirname(__file__), "fonts")
        regular_path = os.path.join(font_dir, "DejaVuSans.ttf")
        bold_path = os.path.join(font_dir, "DejaVuSans-Bold.ttf")

        # Safety: avoid system crash if fonts missing
        for fp in [regular_path, bold_path]:
            if not os.path.exists(fp):
                raise FileNotFoundError(f"❌ Missing required font: {fp}")

        pdf.add_font("DejaVu", "", regular_path, uni=True)
        pdf.add_font("DejaVu", "B", bold_path, uni=True)

        pdf.add_page()
        pdf.set_font("DejaVu", "", 13)
        pdf.set_text_color(*COLORS["text_dark"])

        return pdf


    # ------------------------------------------------------------
    # SECTION TITLE (premium + color accent)
    # ------------------------------------------------------------
    def _section_title(self, pdf, text):
        pdf.ln(4)
        pdf.set_font("DejaVu", "B", 15)
        pdf.set_text_color(*COLORS["primary"])
        pdf.cell(0, 10, f"▌ {text}", ln=True)

        # Reset default text style
        pdf.set_text_color(*COLORS["text_dark"])
        pdf.set_font("DejaVu", "", 12)
        pdf.ln(2)


    # ------------------------------------------------------------
    # Premium dividers
    # ------------------------------------------------------------
    def _divider(self, pdf):
        pdf.set_draw_color(*COLORS["border_light"])
        pdf.set_line_width(0.6)
        y = pdf.get_y()
        pdf.line(10, y, 200, y)
        pdf.ln(5)

    def _subtle_divider(self, pdf):
        pdf.set_draw_color(*COLORS["border_subtle"])
        pdf.set_line_width(0.3)
        y = pdf.get_y()
        pdf.line(10, y, 200, y)
        pdf.ln(3)

    # ------------------------------------------------------------
    # Premium table renderer (alternating colors + alignment)
    # ------------------------------------------------------------
    def _table_header(self, pdf, titles, widths):
        pdf.set_font("DejaVu", "B", 11)
        pdf.set_fill_color(*COLORS["accent"])
        pdf.set_text_color(255, 255, 255)

        for t, w in zip(titles, widths):
            pdf.cell(w, 8, str(t), border=1, align="C", fill=True)

        pdf.ln()
        pdf.set_text_color(*COLORS["text_dark"])


    def _table_row(self, pdf, values, widths, alt=False):
        pdf.set_font("DejaVu", "", 11)

        # Alternating row color
        fill_color = COLORS["table_alt"] if alt else (255, 255, 255)
        pdf.set_fill_color(*fill_color)

        for v, w in zip(values, widths):
            v = "" if v is None else str(v)
            align = "R" if v.replace(".", "", 1).isdigit() else "L"

            pdf.cell(w, 8, v, border=1, fill=True, align=align)

        pdf.ln()


    # ------------------------------------------------------------
    # Safe image insertion (no crashes)
    # ------------------------------------------------------------
    def _insert_image(self, pdf, img_bytes=None, path=None, height=70):
        try:
            if img_bytes:
                tmp_path = os.path.join(
                    tempfile.gettempdir(),
                    f"img_{datetime.now().timestamp()}.png"
                )
                with open(tmp_path, "wb") as f:
                    f.write(img_bytes)

                pdf.image(tmp_path, h=height)
                return

            if path and os.path.exists(path):
                pdf.image(path, h=height)

        except Exception as e:
            pdf.set_text_color(*COLORS["danger"])
            pdf.multi_cell(0, 7, f"⚠ Image could not be inserted.\n{e}")
            pdf.set_text_color(*COLORS["text_dark"])

     
    
    # ------------------------------------------------------------
    # 2. SMART ENRICHMENT ENGINE (Hybrid Matching – Option C)
    # ------------------------------------------------------------
    def get_enriched_sales_df(self, df_sales):
        """
        SMART SALES ENRICHMENT ENGINE (Hybrid Option C)
        ------------------------------------------------
        - Exact match → Token match → Fuzzy match
        - Keeps original item names (NO destructive cleaning)
        - Merges SALES + ITEMS + INVENTORY cleanly
        - Ensures stable output for Excel + Forecasting + Preview
        """

        # ============================================================
        # STEP 0 — BASIC NORMALIZATION
        # ============================================================
        df = df_sales.copy()

        # Detect item column
        possible_item_cols = ["item", "product", "product_name", "name", "sku", "item_name"]
        detected = None
        for c in df.columns:
            if c.lower() in possible_item_cols:
                detected = c
                break

        if detected is None:
            raise Exception("❌ Sales dataset has no item column.")

        df.rename(columns={detected: "item"}, inplace=True)
        df["item"] = df["item"].astype(str).str.strip()

        # Clean numeric fields
        df["quantity"] = pd.to_numeric(df.get("quantity", 0), errors="coerce").fillna(0)

        if "date" in df.columns:
            df["date"] = pd.to_datetime(df["date"], errors="coerce")

        # Normalize for matching only (not replacing original)
        def norm(s: str):
            return (
                str(s)
                .lower()
                .strip()
                .replace("-", " ")
                .replace("_", " ")
            )

        df["item_norm"] = df["item"].apply(norm)

        # ============================================================
        # STEP 1 — LOAD OTHER DATASETS
        # ============================================================
        ds = st.session_state.get("dataset", {})
        df_items = ds.get("items")
        df_inventory = ds.get("inventory")

        # ------------------------------------------------------------
        # prepare metadata datasets for matching
        # ------------------------------------------------------------
        if df_items is not None and not df_items.empty:
            df_items = df_items.copy()
            df_items["item_norm"] = df_items["item"].apply(norm)

        if df_inventory is not None and not df_inventory.empty:
            df_inventory = df_inventory.copy()
            df_inventory["item_norm"] = df_inventory["item"].apply(norm)

        # ============================================================
        # STEP 2 — HYBRID MATCH FUNCTION
        # ============================================================
        from rapidfuzz import process, fuzz

        def hybrid_match(x, pool):
            """Exact → Token → Fuzzy matching"""
            x = str(x).lower().strip()

            if not pool:
                return None

            # ----- 1. Exact match -----
            if x in pool:
                return x

            # ----- 2. Token match ("milk" matches "fresh milk 1l") -----
            for p in pool:
                if x in p or p in x:
                    return p

            # ----- 3. Fuzzy match -----
            best, score, _ = process.extractOne(
                x, pool, scorer=fuzz.token_sort_ratio
            )
            if score >= 80:
                return best

            return None

        # ============================================================
        # STEP 3 — MERGE WITH ITEMS METADATA
        # ============================================================
        if df_items is not None and not df_items.empty:
            item_pool = df_items["item_norm"].tolist()

            df["match_items"] = df["item_norm"].apply(
                lambda x: hybrid_match(x, item_pool)
            )

            df = df.merge(
                df_items.drop(columns=["item"]),
                left_on="match_items",
                right_on="item_norm",
                how="left",
                suffixes=("", "_meta")
            )

        # ============================================================
        # STEP 4 — MERGE WITH INVENTORY
        # ============================================================
        if df_inventory is not None and not df_inventory.empty:
            inv_pool = df_inventory["item_norm"].tolist()

            df["match_inv"] = df["item_norm"].apply(
                lambda x: hybrid_match(x, inv_pool)
            )

            df = df.merge(
                df_inventory.drop(columns=["item"]),
                left_on="match_inv",
                right_on="item_norm",
                how="left",
                suffixes=("", "_inv")
            )

        # ============================================================
        # STEP 5 — BUSINESS SAFETY COLUMNS
        # ============================================================
        defaults = {
            "current_stock": 0,
            "reorder_point": 0,
            "recommended_reorder_qty": 0,
            "status": "Unknown",
            "unit": "Unit",
            "category": "General",
            "price": 0,
        }

        for col, default in defaults.items():
            if col not in df.columns:
                df[col] = default
            else:
                df[col] = df[col].fillna(default)

        # numeric safety
        for col in ["current_stock", "reorder_point", "recommended_reorder_qty", "price"]:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

        # ============================================================
        # STEP 6 — BUSINESS METRICS
        # ============================================================
        df["out_of_stock"] = df["current_stock"] <= 0
        df["low_stock"] = df["current_stock"] < df["reorder_point"]

        df["stock_health_pct"] = (
            (df["current_stock"] / df["reorder_point"].replace(0, 1)) * 100
        ).round(1)

        df["recommended_reorder_qty"] = (
            (df["reorder_point"] - df["current_stock"]).clip(lower=0)
        )

        # ============================================================
        # STEP 7 — CLEAN UP TECHNICAL COLUMNS
        # ============================================================
        df.drop(
            columns=[
                c for c in df.columns
                if c.endswith("_meta") or c.endswith("_inv")
                or c in ["item_norm", "match_items", "match_inv"]
            ],
            inplace=True,
            errors="ignore"
        )

        # ============================================================
        # STEP 8 — ORDER COLUMNS
        # ============================================================
        preferred = [
            "date", "item", "quantity",
            "category", "unit",
            "current_stock", "reorder_point",
            "recommended_reorder_qty", "status",
            "low_stock", "out_of_stock", "stock_health_pct"
        ]

        ordered = [c for c in preferred if c in df.columns]
        rest = [c for c in df.columns if c not in ordered]

        df = df[ordered + rest]

        # ============================================================
        # STEP 9 — SORT BEFORE RETURNING
        # ============================================================
        if "date" in df.columns:
            df = df.sort_values(["date", "item"])

        return df.reset_index(drop=True)

    
    # ------------------------------------------------------------
    # 3. PREMIUM CLEAN EXCEL EXPORT (Smart + Crash-Proof)
    # ------------------------------------------------------------
    def export_excel_bytes(self, df, filename="report.xlsx"):
        """
        ENTERPRISE-GRADE EXCEL EXPORTER
        - Auto-adapts to any dataframe
        - Fully safe: no chart, pivot, or formatting crash
        - Includes smart summaries, pivots, conditional formatting
        """

        output = io.BytesIO()
        df = df.copy()

        # ============================================================
        # 1. SMART COLUMN ORDER
        # ============================================================
        preferred = [
            "date", "item", "quantity",
            "category", "unit", "price",
            "current_stock", "reorder_point",
            "recommended_reorder_qty", "status",
            "low_stock", "out_of_stock", "stock_health_pct",
        ]

        ordered = [c for c in preferred if c in df.columns]
        rest = [c for c in df.columns if c not in ordered]
        df = df[ordered + rest]

        # ============================================================
        # 2. WRITE EXCEL FILE
        # ============================================================
        with pd.ExcelWriter(output, engine="openpyxl") as writer:

            # ---------------- DATA SHEET ----------------
            df.to_excel(writer, sheet_name="Data", index=False)
            data_sheet = writer.sheets["Data"]

            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
            from openpyxl.chart import BarChart, LineChart, Reference

            # =================== HEADER STYLE ===================
            header_fill = PatternFill(start_color="2A5BD8", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            thin = Side(style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for cell in data_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # =================== AUTO WIDTH ===================
            for col_cells in data_sheet.columns:
                values = [str(c.value) for c in col_cells if c.value is not None]
                max_len = max((len(v) for v in values), default=8)
                data_sheet.column_dimensions[col_cells[0].column_letter].width = max_len + 2

            data_sheet.freeze_panes = "A2"
            last_row = df.shape[0] + 1

            # ============================================================
            # 3. CONDITIONAL FORMATTING (Safely Applied)
            # ============================================================
            for idx, col in enumerate(df.columns, start=1):
                col_letter = data_sheet.cell(row=1, column=idx).column_letter

                # Highlight missing values (NaN)
                data_sheet.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{last_row}",
                    CellIsRule(
                        operator="equal",
                        formula=["\"\""],
                        fill=PatternFill(start_color="FFCCCC", fill_type="solid"),
                    ),
                )

            # Quantity highlight
            if "quantity" in df.columns:
                q_idx = df.columns.get_loc("quantity") + 1
                q_letter = data_sheet.cell(row=1, column=q_idx).column_letter
                data_sheet.conditional_formatting.add(
                    f"{q_letter}2:{q_letter}{last_row}",
                    CellIsRule(
                        operator="greaterThan",
                        formula=["100"],
                        fill=PatternFill(start_color="FFF2CC", fill_type="solid"),
                    ),
                )

            # Stock heatmap
            if "current_stock" in df.columns:
                cs_idx = df.columns.get_loc("current_stock") + 1
                cs_letter = data_sheet.cell(row=1, column=cs_idx).column_letter
                data_sheet.conditional_formatting.add(
                    f"{cs_letter}2:{cs_letter}{last_row}",
                    ColorScaleRule(
                        start_type="min", start_color="FF0000",
                        mid_type="percentile", mid_value=50, mid_color="FFFF00",
                        end_type="max", end_color="00B050"
                    )
                )

            # ============================================================
            # 4. PIVOT SHEETS (Only if possible)
            # ============================================================
            pivot_sheets = {}

            # ---- Pivot 1: Sales Summary ----
            if "item" in df.columns and "quantity" in df.columns:
                pivot1 = (
                    df.groupby("item")["quantity"]
                    .agg(["sum", "mean", "min", "max", "std"])
                    .reset_index()
                )
                pivot1.columns = ["Item", "Total", "Avg", "Min", "Max", "StdDev"]
                pivot1.to_excel(writer, sheet_name="Pivot_Sales", index=False)
                pivot_sheets["pivot_sales"] = writer.sheets["Pivot_Sales"]

            # ---- Pivot 2: Inventory ----
            if "item" in df.columns and "current_stock" in df.columns:
                inv_cols = [
                    c for c in
                    ["item", "current_stock", "reorder_point", "status"]
                    if c in df.columns
                ]
                pivot2 = df[inv_cols].drop_duplicates()
                pivot2.to_excel(writer, sheet_name="Pivot_Inventory", index=False)
                pivot_sheets["pivot_inventory"] = writer.sheets["Pivot_Inventory"]

            # ============================================================
            # 5. CHARTS (Only if pivots exist)
            # ============================================================
            chart_sheet = writer.book.create_sheet("Charts")

            # -------- Bar Chart: Sales --------
            if "pivot_sales" in pivot_sheets:
                p = pivot_sheets["pivot_sales"]
                rows = p.max_row

                chart = BarChart()
                chart.title = "Total Sales per Item"
                chart.y_axis.title = "Quantity"
                chart.x_axis.title = "Items"

                data_ref = Reference(p, min_col=2, min_row=1, max_row=rows)
                cats_ref = Reference(p, min_col=1, min_row=2, max_row=rows)

                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)

                chart_sheet.add_chart(chart, "A1")

            # -------- Line Chart: Daily Trend --------
            if "date" in df.columns and "quantity" in df.columns:
                trend = df.groupby("date")["quantity"].sum().reset_index()
                trend.to_excel(writer, sheet_name="TrendData", index=False)
                td = writer.sheets["TrendData"]

                rows = td.max_row
                line = LineChart()
                line.title = "Daily Quantity Trend"

                data_ref = Reference(td, min_col=2, min_row=1, max_row=rows)
                cats_ref = Reference(td, min_col=1, min_row=2, max_row=rows)

                line.add_data(data_ref, titles_from_data=True)
                line.set_categories(cats_ref)

                chart_sheet.add_chart(line, "A20")

            # ============================================================
            # 6. SUMMARY SHEET (Clean & Informative)
            # ============================================================
            summary = writer.book.create_sheet("Summary")
            summary["A1"] = "Export Summary"
            summary["A1"].font = Font(bold=True, size=14)

            summary["A3"] = "Generated On"
            summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")

            summary["A4"] = "Total Rows"
            summary["B4"] = len(df)

            summary["A5"] = "Total Columns"
            summary["B5"] = len(df.columns)

            summary["A7"] = "Missing Values:"
            for i, (col, missing) in enumerate(df.isna().sum().items(), start=8):
                summary[f"A{i}"] = col
                summary[f"B{i}"] = missing

        return output.getvalue(), filename

    

    
    # ------------------------------------------------------------
    # 4. PREMIUM INVENTORY SUMMARY PDF (Enterprise Version)
    # ------------------------------------------------------------
    def export_inventory_pdf_bytes(self, df_inventory):

        # ---- SAFETY COPY ----
        df = df_inventory.copy()

        # ---- Ensure essential columns exist ----
        for col in ["item", "current_stock", "reorder_point", "recommended_reorder_qty", "status"]:
            if col not in df.columns:
                if col in ["current_stock", "reorder_point", "recommended_reorder_qty"]:
                    df[col] = 0
                else:
                    df[col] = "Unknown"

        # ---- Ensure numeric columns are numeric ----
        for col in ["current_stock", "reorder_point", "recommended_reorder_qty"]:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

        pdf = self._new_pdf()

        # ---------------- TITLE ----------------
        self._section_title(pdf, "Inventory Summary Report")
        self._divider(pdf)

        # ---------------- META SUMMARY ----------------
        total_items = len(df)
        low_stock = (df["current_stock"] < df["reorder_point"]).sum()
        out_stock = (df["current_stock"] == 0).sum()
        avg_rop = df["reorder_point"].mean()
        percent_low = round((low_stock / total_items) * 100, 1) if total_items else 0

        pdf.set_font("DejaVu", "", 12)
        pdf.multi_cell(
            0, 8,
            f"Total Unique Items: {total_items}\n"
            f"Low Stock Items: {low_stock} ({percent_low}%)\n"
            f"Out-of-Stock Items: {out_stock}\n"
            f"Average ROP: {avg_rop:.1f}"
        )
        pdf.ln(3)

        # ---------------- STOCK HEALTH BADGE ----------------
        pdf.set_font("DejaVu", "B", 12)

        total_stock = df["current_stock"].sum()
        total_rop = df["reorder_point"].sum()

        if total_rop > 0:
            health_score = round((total_stock / total_rop) * 100, 1)
        else:
            health_score = 100

        if health_score < 60:
            color = COLORS["danger"]; label = "Poor Stock Health"; icon = "🔴"
        elif health_score < 85:
            color = COLORS["accent"]; label = "Moderate Stock Health"; icon = "🟡"
        else:
            color = COLORS["success"]; label = "Excellent Stock Health"; icon = "🟢"

        pdf.set_text_color(*color)
        pdf.multi_cell(0, 8, f"{icon} Stock Health Score: {health_score}% — {label}")
        pdf.set_text_color(*COLORS["text_dark"])
        pdf.ln(5)

        # ---------------- INVENTORY TABLE ----------------
        self._section_title(pdf, "Inventory Table")
        self._subtle_divider(pdf)

        headers = ["Item", "Stock", "ROP", "Reorder Qty", "Status"]
        widths = [60, 25, 25, 35, 40]

        self._table_header(pdf, headers, widths)

        alt = False
        for _, row in df.iterrows():

            # color coding
            if row["current_stock"] == 0:
                pdf.set_text_color(*COLORS["danger"])
            elif row["current_stock"] < row["reorder_point"]:
                pdf.set_text_color(255, 165, 0)
            else:
                pdf.set_text_color(*COLORS["text_dark"])

            vals = [
                row["item"],
                row["current_stock"],
                row["reorder_point"],
                row["recommended_reorder_qty"],
                row["status"]
            ]

            self._table_row(pdf, vals, widths, alt=alt)
            alt = not alt

            pdf.set_text_color(*COLORS["text_dark"])

        # ---------------- INSIGHTS ----------------
        self._section_title(pdf, "Insights & Analysis")
        self._divider(pdf)
        pdf.set_font("DejaVu", "", 11)

        low_items = df.sort_values("current_stock").head(5)
        pdf.multi_cell(0, 7, "🔻 Top 5 Lowest Stock Items:")
        for _, r in low_items.iterrows():
            pdf.cell(0, 7, f"- {r['item']} (Stock: {r['current_stock']})", ln=True)

        pdf.ln(3)

        high_rop = df.sort_values("reorder_point", ascending=False).head(5)
        pdf.multi_cell(0, 7, "📌 Highest Reorder Points:")
        for _, r in high_rop.iterrows():
            pdf.cell(0, 7, f"- {r['item']} (ROP: {r['reorder_point']})", ln=True)

        pdf.ln(3)

        critical = df[df["current_stock"] == 0]
        if not critical.empty:
            pdf.set_text_color(*COLORS["danger"])
            pdf.multi_cell(0, 8, "⚠ Out-of-Stock Items:")
            pdf.set_text_color(*COLORS["text_dark"])
            for _, r in critical.iterrows():
                pdf.cell(0, 7, f"- {r['item']}", ln=True)

        # ---------------- ABC ANALYSIS ----------------
        pdf.ln(4)
        self._section_title(pdf, "ABC Inventory Classification")
        self._subtle_divider(pdf)

        df["weight"] = df["current_stock"] * df["reorder_point"]
        df_sorted = df.sort_values("weight", ascending=False)

        df_sorted["share"] = df_sorted["weight"] / (df_sorted["weight"].sum() + 1)
        df_sorted["cum"] = df_sorted["share"].cumsum()

        df_sorted["abc"] = df_sorted["cum"].apply(
            lambda x: "A" if x <= 0.70 else ("B" if x <= 0.90 else "C")
        )

        pdf.multi_cell(
            0, 7,
            f"A-items (High impact): {(df_sorted['abc'] == 'A').sum()}\n"
            f"B-items (Medium impact): {(df_sorted['abc'] == 'B').sum()}\n"
            f"C-items (Low impact): {(df_sorted['abc'] == 'C').sum()}"
        )

        pdf_bytes = pdf.output(dest="S")
        return bytes(pdf_bytes), "inventory_summary.pdf"


   
    
    # ------------------------------------------------------------
    # 5. LEVEL 3 – ADVANCED FORECAST REPORT PDF (Hybrid Upgrade)
    # ------------------------------------------------------------
    def export_forecast_pdf_bytes(self, df_sales, df_inventory, item_name, days=30):

        # -------------------------------
        # 1) Call forecasting engine
        # -------------------------------
        tools = AgentTools(df_sales, df_inventory)
        result = tools.forecast_item(item_name, days=days)

        history = result.get("history", [])
        df_hist = pd.DataFrame(history)

        forecast = result.get("forecast", [])
        df_fc = pd.DataFrame(forecast)

        chart_b64 = result.get("chart_data")  # may be None

        # ------------------------------------------
        # 2) UNIVERSAL SANITIZERS (CRITICAL UPGRADE)
        # ------------------------------------------

        # -- Fix missing quantity --
        qty_candidates = ["quantity", "y", "value"]
        found = None
        for q in qty_candidates:
            if q in df_hist.columns:
                found = q
                break

        if found is None:
            df_hist["quantity"] = 0
        else:
            df_hist["quantity"] = pd.to_numeric(df_hist[found], errors="coerce").fillna(0)

        # -- Fix missing date --
        if "date" not in df_hist.columns:
            if "ds" in df_hist.columns:
                df_hist["date"] = df_hist["ds"]
            else:
                df_hist["date"] = pd.NaT

        df_hist["date"] = pd.to_datetime(df_hist["date"], errors="coerce")

        # -- Forecast sanitizers --
        if "ds" in df_fc.columns:
            df_fc["ds"] = pd.to_datetime(df_fc["ds"], errors="coerce")
        if "yhat" in df_fc.columns:
            df_fc["yhat"] = pd.to_numeric(df_fc["yhat"], errors="coerce").fillna(0)
        else:
            df_fc["yhat"] = 0

        # ------------------------------------------
        # 3) Start PDF
        # ------------------------------------------
        pdf = self._new_pdf()

        self._section_title(pdf, f"Forecast Report — {item_name}")
        self._divider(pdf)
        pdf.set_font("DejaVu", size=12)

        # ------------------------------------------
        # 4) Insights (Stable + Safe)
        # ------------------------------------------

        # Growth Trend
        if len(df_hist) > 1:
            first = df_hist["quantity"].iloc[0]
            last = df_hist["quantity"].iloc[-1]
            growth_pct = ((last - first) / max(first, 1)) * 100
        else:
            growth_pct = 0

        if growth_pct > 25:
            trend = "📈 Strong Upward Trend"
        elif growth_pct > 10:
            trend = "↗️ Moderate Increase"
        elif abs(growth_pct) < 10:
            trend = "➖ Stable Demand"
        else:
            trend = "📉 Declining Demand"

        # Volatility
        vol = float(df_hist["quantity"].std()) if len(df_hist) > 2 else 0
        if vol < 3:
            vol_label = "🟢 Very Stable"
        elif vol < 8:
            vol_label = "🟡 Moderate Fluctuation"
        else:
            vol_label = "🔴 Highly Volatile"

        # Forecast confidence
        spread = df_fc["yhat"].std() if len(df_fc) > 1 else None
        if spread is None:
            conf = "N/A"
        elif spread < 2:
            conf = "🟢 High Confidence"
        elif spread < 6:
            conf = "🟡 Medium Confidence"
        else:
            conf = "🔴 Low Confidence"

        # Peak day (SAFE)
        if not df_fc.empty:
            peak_idx = df_fc["yhat"].idxmax()
            peak_row = df_fc.loc[peak_idx]

            peak_day = (
                str(peak_row["ds"].date())
                if isinstance(peak_row["ds"], pd.Timestamp)
                else "-"
            )

            try:
                peak_value = float(peak_row["yhat"])
            except:
                peak_value = None
        else:
            peak_day, peak_value = "-", None

        peak_value_str = f"{peak_value:.2f}" if isinstance(peak_value, (int, float)) else "N/A"

        # Weekly patterns
        if len(df_hist) > 14 and df_hist["date"].notna().sum() > 14:
            weekday_avg = df_hist.groupby(df_hist["date"].dt.day_name())["quantity"].mean()
            best_day = weekday_avg.idxmax()
            worst_day = weekday_avg.idxmin()
            weekly_pattern = f"Best Day: **{best_day}**, Weakest: **{worst_day}**."
        else:
            weekly_pattern = "Not enough data."

        # Anomalies
        q1 = df_hist["quantity"].quantile(0.25)
        q3 = df_hist["quantity"].quantile(0.75)
        iqr = q3 - q1
        spikes = df_hist[df_hist["quantity"] > q3 + 2.2 * iqr]
        spike_msg = f"{len(spikes)} anomalies detected." if len(spikes) > 0 else "No anomalies detected."

        # Demand class
        avg_qty = df_hist["quantity"].mean()
        if avg_qty < 5:
            demand_class = "🟦 Low Demand SKU"
        elif avg_qty < 12:
            demand_class = "🟧 Moderate Demand SKU"
        else:
            demand_class = "🟥 High Demand SKU"

        # ------------------------------------------
        # 5) Insights Block
        # ------------------------------------------
        pdf.multi_cell(
            0, 8,
            f"📌 **Forecast Horizon:** {days} days\n"
            f"📊 **Trend:** {trend}\n"
            f"📉 **Volatility:** {vol_label} (σ = {vol:.2f})\n"
            f"🎯 **Demand Classification:** {demand_class}\n"
            f"🔍 **Forecast Confidence:** {conf}\n"
            f"🏁 **Peak Forecast Day:** {peak_day} ({peak_value_str})\n"
            f"📆 **Weekly Pattern:** {weekly_pattern}\n"
            f"⚡ **Anomaly Check:** {spike_msg}\n"
        )

        pdf.ln(6)

        # ------------------------------------------
        # 6) Chart Section (Guaranteed Safe)
        # ------------------------------------------
        self._section_title(pdf, "Forecast Chart")
        self._subtle_divider(pdf)

        if chart_b64:
            try:
                import base64, tempfile, os
                img_bytes = base64.b64decode(chart_b64)
                temp_path = os.path.join(tempfile.gettempdir(), "forecast_chart.png")
                with open(temp_path, "wb") as f:
                    f.write(img_bytes)
                pdf.image(temp_path, w=180)
            except Exception:
                pdf.multi_cell(0, 8, "Chart could not be inserted.")
        else:
            pdf.multi_cell(0, 8, "No chart available.")

        pdf.ln(6)

        # ------------------------------------------
        # 7) Forecast Table
        # ------------------------------------------
        self._section_title(pdf, "Top 10 Forecasted Days")
        self._subtle_divider(pdf)

        if not df_fc.empty:
            df_fc_sorted = df_fc.sort_values("yhat", ascending=False).head(10)

            headers = ["Date", "Prediction", "Trend vs Today"]
            widths = [55, 45, 70]
            self._table_header(pdf, headers, widths)

            last_hist = df_hist["quantity"].iloc[-1] if not df_hist.empty else 0

            alt = False
            for _, row in df_fc_sorted.iterrows():
                d = (
                    str(row["ds"].date())
                    if isinstance(row["ds"], pd.Timestamp)
                    else "-"
                )
                pred = f"{float(row['yhat']):.2f}"
                trend_arrow = "↑ Higher" if row["yhat"] > last_hist else "↓ Lower"

                vals = [d, pred, trend_arrow]
                self._table_row(pdf, vals, widths, alt=alt)
                alt = not alt

        # ------------------------------------------
        # 8) Footer
        # ------------------------------------------
        pdf.ln(6)
        pdf.set_font("DejaVu", size=11)
        pdf.multi_cell(
            0, 8,
            "📝 **Notes:**\n"
            "- This forecast blends statistical + machine learning models.\n"
            "- Volatility and anomalies indicate reliability of predictions.\n"
            "- Weekly pattern detection is based on historical sales.\n"
            "- For deeper AI explanations, use the AI Assistant in the dashboard."
        )

        pdf_bytes = pdf.output(dest="S")
        return bytes(pdf_bytes), f"forecast_{item_name}.pdf"



    #
    # ------------------------------------------------------------
    # 6. PREMIUM REORDER RISK REPORT PDF (Enterprise Version)
    # ------------------------------------------------------------
    def export_reorder_pdf_bytes(self, df_sales, df_inventory):
        tools = AgentTools(df_sales, df_inventory)
        predictor = MLPredictor()

        pdf = self._new_pdf()

        # ======================
        # TITLE
        # ======================
        self._section_title(pdf, "Reorder Risk Analysis")
        self._divider(pdf)
        pdf.set_font("DejaVu", size=12)

        # ======================
        # COL SAFETY
        # ======================
        df_inv = df_inventory.copy()
        for col in ["item", "current_stock", "reorder_point"]:
            if col not in df_inv.columns:
                df_inv[col] = 0

        # ======================
        # RISK PROCESSING
        # ======================
        risk_rows = []
        risk_stats = {"critical": 0, "high": 0, "medium": 0, "low": 0, "safe": 0}

        for item in df_inv["item"].unique():

            valid, df_item = tools._validate_item(item)
            if not valid:
                continue

            # Clean & predict
            df_clean = tools.cleaner.clean_sales(df_item)
            reorder_result = predictor.predict_reorder(item, df_clean)

            # Expected keys:
            # reorder_probability, recommend_reorder
            # days_until_stockout, recommended_qty
            prob = reorder_result.get("reorder_probability", 0)
            days_left = reorder_result.get("days_until_stockout")
            rec_qty = reorder_result.get("recommended_qty", 0)

            # Classify probability into color-coded risk
            if days_left == 0:
                risk_label = "Critical"
                risk_stats["critical"] += 1
            elif prob > 0.80:
                risk_label = "High"
                risk_stats["high"] += 1
            elif prob > 0.55:
                risk_label = "Medium"
                risk_stats["medium"] += 1
            elif prob > 0.35:
                risk_label = "Low"
                risk_stats["low"] += 1
            else:
                risk_label = "Safe"
                risk_stats["safe"] += 1

            # Compute projected stockout date
            if days_left is not None:
                try:
                    projected = (pd.Timestamp.today() + pd.Timedelta(days=days_left)).date()
                except:
                    projected = "-"
            else:
                projected = "-"

            risk_rows.append({
                "item": item,
                "prob": round(prob, 3),
                "risk": risk_label,
                "days_left": days_left,
                "projected": projected,
                "recommended_qty": rec_qty
            })

        # ======================
        # SUMMARY BLOCK
        # ======================
        pdf.multi_cell(
            0, 8,
            f"Total Items Analyzed: {len(risk_rows)}\n"
            f"Critical: {risk_stats['critical']}\n"
            f"High Risk: {risk_stats['high']}\n"
            f"Medium Risk: {risk_stats['medium']}\n"
            f"Low Risk: {risk_stats['low']}\n"
            f"Safe: {risk_stats['safe']}\n"
        )
        pdf.ln(4)

        # ======================
        # CRITICAL ALERTS
        # ======================
        critical_items = [r for r in risk_rows if r["risk"] == "Critical"]
        urgent_items = [r for r in risk_rows if r["days_left"] is not None and r["days_left"] <= 3]

        self._section_title(pdf, "Critical Alerts")
        self._divider(pdf)

        if not critical_items:
            pdf.multi_cell(0, 7, "No items completely out of stock.")
        else:
            pdf.multi_cell(0, 7, "❗ Out-of-Stock Items:")
            for r in critical_items:
                pdf.cell(
                    0, 7,
                    f"- {r['item']} (Recommended: {r['recommended_qty']})",
                    ln=True
                )

        pdf.ln(3)

        if not urgent_items:
            pdf.multi_cell(0, 7, "No items are within 3 days of stockout.")
        else:
            pdf.multi_cell(0, 7, "⚠️ Urgent Reorder Needed (≤ 3 days left):")
            for r in urgent_items:
                pdf.cell(
                    0, 7,
                    f"- {r['item']} ({r['days_left']} days left, Qty: {r['recommended_qty']})",
                    ln=True
                )

        pdf.ln(5)

        # ======================
        # MAIN RISK TABLE
        # ======================
        self._section_title(pdf, "Detailed Reorder Risk Table")
        self._subtle_divider(pdf)

        headers = ["Item", "Risk", "Prob.", "Days Left", "Stockout", "Reorder Qty"]
        widths = [55, 25, 25, 25, 35, 35]
        self._table_header(pdf, headers, widths)

        alt = False
        for r in risk_rows:

            # color by risk
            risk = r["risk"].lower()
            if risk == "critical":
                pdf.set_text_color(200, 0, 0)
            elif risk == "high":
                pdf.set_text_color(255, 120, 0)
            elif risk == "medium":
                pdf.set_text_color(240, 200, 0)
            elif risk == "low":
                pdf.set_text_color(0, 150, 0)
            else:
                pdf.set_text_color(80, 80, 80)

            vals = [
                r["item"],
                r["risk"],
                r["prob"],
                r["days_left"],
                str(r["projected"]),
                r["recommended_qty"]
            ]

            self._table_row(pdf, vals, widths, alt=alt)
            alt = not alt
            pdf.set_text_color(*COLORS["text_dark"])

        pdf.ln(5)

        # ======================
        # TOP 5 MOST URGENT
        # ======================
        self._section_title(pdf, "Top 5 Most Urgent Items")
        self._divider(pdf)

        sorted_by_urgency = sorted(
            risk_rows,
            key=lambda x: x["days_left"] if x["days_left"] is not None else 999
        )[:5]

        if not sorted_by_urgency:
            pdf.multi_cell(0, 7, "No urgent items detected.")
        else:
            for r in sorted_by_urgency:
                pdf.cell(
                    0, 7,
                    f"- {r['item']} ({r['days_left']} days left, Qty: {r['recommended_qty']})",
                    ln=True
                )

        pdf.ln(5)

        # ======================
        # FINAL NOTES
        # ======================
        pdf.set_font("DejaVu", size=11)
        pdf.multi_cell(
            0, 8,
            "📝 Notes:\n"
            "- Risk scores are predicted using ML classifiers.\n"
            "- Stockout days are based on recent sales velocity.\n"
            "- Recommended quantities aim to prevent repeat stockouts.\n"
            "- Monitor Medium-risk items daily.\n"
            "- Safe items require only periodic review."
        )

        return bytes(pdf.output(dest="S")), "reorder_risk_report.pdf"


    
    # ------------------------------------------------------------
    # 7. ENTERPRISE VOLATILITY REPORT (Excel with advanced insights)
    # ------------------------------------------------------------
    def export_volatility_excel_bytes(self, df_sales):
        tools = AgentTools(df_sales, df_sales)

        rows = []
        for item in sorted(df_sales["item"].unique()):
            result = tools.get_item_volatility(item)

            if not isinstance(result, dict) or result.get("error"):
                continue

            std_val = result.get("std_dev", 0)
            mean_val = result.get("mean_quantity", 0)
            cv = std_val / mean_val if mean_val > 0 else None

            # Volatility Class
            if cv is None:
                vol_class = "Unknown"
            elif cv < 0.10:
                vol_class = "Very Stable"
            elif cv < 0.25:
                vol_class = "Stable"
            elif cv < 0.40:
                vol_class = "Volatile"
            else:
                vol_class = "Highly Volatile"

            # Volatility Score (normalized 0–100)
            vol_score = round((cv or 0) * 100, 1)

            # Trend arrows (strong vs weak)
            history = result.get("history", [])
            trend_sig = ""
            if len(history) > 2:
                for a, b in zip(history[:-1], history[1:]):
                    diff = b - a
                    if diff > 3:
                        trend_sig += "↑"
                    elif diff > 0:
                        trend_sig += "↗"
                    elif diff == 0:
                        trend_sig += "→"
                    elif diff < -3:
                        trend_sig += "↓"
                    else:
                        trend_sig += "↘"

            rows.append({
                "item": item,
                "std_dev": std_val,
                "mean_qty": mean_val,
                "min_qty": result.get("min_quantity"),
                "max_qty": result.get("max_quantity"),
                "window": result.get("window"),
                "CV": round(cv, 3) if cv else None,
                "Volatility Score": vol_score,
                "Volatility Class": vol_class,
                "trend_signal": trend_sig,
                "history": history
            })

        df_vol = pd.DataFrame(rows)

        # ---------------- Excel Writing ----------------
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_vol.drop(columns=["history"]).to_excel(
                writer,
                index=False,
                sheet_name="Volatility"
            )

            workbook = writer.book
            sheet = writer.sheets["Volatility"]

            # Freeze header
            sheet.freeze_panes = "A2"

            # Auto width
            for col in sheet.columns:
                max_len = max(len(str(c.value)) for c in col)
                sheet.column_dimensions[col[0].column_letter].width = max_len + 3

            # ---------------- CONDITIONAL FORMATTING ----------------
            from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

            # Heatmap: CV (G column)
            cv_col = "G"
            sheet.conditional_formatting.add(
                f"{cv_col}2:{cv_col}{len(df_vol) + 1}",
                ColorScaleRule(
                    start_type="min", start_color="00A6FF00",   # green
                    mid_type="percentile", mid_value=50, mid_color="00FFD966",  # yellow
                    end_type="max", end_color="00FF0000"       # red
                )
            )

            # Highlight Highly Volatile
            vol_class_col = df_vol.columns.get_loc("Volatility Class") + 1
            letter = chr(ord("A") + vol_class_col - 1)
            sheet.conditional_formatting.add(
                f"{letter}2:{letter}{len(df_vol)+1}",
                CellIsRule(operator="equal", formula=['"Highly Volatile"'],
                        fill=PatternFill(start_color="FF6666", fill_type="solid"))
            )

            # ---------------- TREND COLUMN (advanced arrows) ----------------
            trend_col_letter = chr(ord("A") + df_vol.shape[1])  # next empty column
            trend_col_index = df_vol.shape[1] + 1

            sheet.cell(row=1, column=trend_col_index).value = "Trend"
            sheet.cell(row=1, column=trend_col_index).font = Font(bold=True)

            for i, row in enumerate(rows, start=2):
                trend = row.get("trend_signal", "")
                sheet.cell(row=i, column=trend_col_index, value=trend if trend else "N/A")

            sheet.column_dimensions[trend_col_letter].width = 18

            # ---------------- INSIGHTS SHEET ----------------
            insights = workbook.create_sheet("Insights")

            def write_block(title, lines, row_start):
                insights.cell(row=row_start, column=1, value=title)
                insights.cell(row=row_start, column=1).font = Font(bold=True)
                row = row_start + 1
                for ln in lines:
                    insights.cell(row=row, column=1, value=ln)
                    row += 1
                return row + 1

            # Top 5 volatile
            top_volatile = df_vol.sort_values("CV", ascending=False).head(5)
            r = write_block("Top 5 Highly Volatile Items:",
                            [f"{x.item} (CV={x.CV})" for x in top_volatile.itertuples()],
                            1)

            # Top 5 stable
            top_stable = df_vol.sort_values("CV", ascending=True).head(5)
            r = write_block("Top 5 Most Stable Items:",
                            [f"{x.item} (CV={x.CV})" for x in top_stable.itertuples()],
                            r)

            # Items with sharp spikes
            spike_items = df_vol[df_vol["trend_signal"].str.contains("↑↑|↓↓", na=False)]
            r = write_block("Items With Sharp Volatility Spikes:",
                            [f"{x.item} ({x.trend_signal})" for x in spike_items.itertuples()],
                            r)

            # Longest window
            long_window = df_vol.sort_values("window", ascending=False).head(5)
            r = write_block("Longest Window Size (Better Forecasting Reliability):",
                            [f"{x.item} (Window={x.window})" for x in long_window.itertuples()],
                            r)

            # Recommendations
            write_block(
                "Recommended Actions:",
                [
                    "Highly Volatile: Increase safety stock + monitor daily.",
                    "Volatile: Review reorder logic weekly.",
                    "Stable: Use standard cycle-based replenishment.",
                    "Very Stable: Ideal for long-term forecasting optimization.",
                    "Sharp Spike Items: Investigate causes (promo, supply, seasonality)."
                ],
                r
            )

        return output.getvalue(), "volatility_report.xlsx"
